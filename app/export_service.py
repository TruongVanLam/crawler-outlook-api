"""
Export service for generating Excel files and ZIP archives
"""
import io
import zipfile
import pandas as pd
from datetime import datetime
from typing import List, Dict, Any
from sqlalchemy.orm import Session
from sqlalchemy import and_

from models import Email, Account, MetaReceipt
from .email_utils_bs4 import extract_meta_receipt_info_combined


class ExportService:
    """Service for exporting email data to Excel files"""
    
    def __init__(self, db: Session):
        self.db = db
    
    def get_meta_receipt_emails(
        self, 
        account_ids: List[int], 
        from_date: str, 
        to_date: str
    ) -> Dict[int, List[Dict[str, Any]]]:
        """
        Lấy tất cả Meta receipts từ bảng meta_receipts theo account_ids và khoảng thời gian
        """
        receipts_by_account = {}
        
        for account_id in account_ids:
            # Lấy meta receipts từ database theo điều kiện
            receipts = self.db.query(MetaReceipt).filter(
                and_(
                    MetaReceipt.account_id == account_id,
                    MetaReceipt.date >= f"{from_date}T00:00:00",
                    MetaReceipt.date <= f"{to_date}T23:59:59"
                )
            ).order_by(MetaReceipt.date.desc()).all()
            
            account_receipts = []
            
            for receipt in receipts:
                # Tạo row data từ meta receipt
                row_data = {
                    'Date': receipt.date.isoformat() if receipt.date else None,
                    'Account ID': receipt.account_id_meta,
                    'Transaction ID': receipt.transaction_id,
                    'Payment': receipt.payment,
                    'Card Number': f"Visa · {receipt.card_number}" if receipt.card_number else None,
                    'Reference Number': receipt.reference_number,
                    'Status': receipt.status
                }
                account_receipts.append(row_data)
            
            receipts_by_account[account_id] = account_receipts
        
        return receipts_by_account
    
    def create_excel_file(self, data: List[Dict[str, Any]], account_email: str) -> io.BytesIO:
        """
        Tạo file Excel từ data với borders và column width tối ưu
        """
        df = pd.DataFrame(data)
        
        # Đảm bảo có đủ các cột theo format yêu cầu
        required_columns = ['Date', 'Account ID', 'Transaction ID', 'Payment', 'Card Number', 'Reference Number', 'Status']
        for col in required_columns:
            if col not in df.columns:
                df[col] = None
        
        # Sắp xếp theo thứ tự cột yêu cầu
        df = df[required_columns]
        
        # Tạo Excel file
        output = io.BytesIO()
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            df.to_excel(writer, index=False, sheet_name='Meta Receipts')
            
            # Lấy workbook và worksheet để styling
            workbook = writer.book
            worksheet = writer.sheets['Meta Receipts']
            
            # Import openpyxl styles
            from openpyxl.styles import Border, Side, Alignment, Font
            from openpyxl.utils import get_column_letter
            
            # Định nghĩa border style
            thin_border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            
            # Định nghĩa header style
            header_font = Font(bold=True)
            header_alignment = Alignment(horizontal='center', vertical='center')
            
            # Định nghĩa cell alignment
            cell_alignment = Alignment(horizontal='left', vertical='center')
            
            # Styling cho header row (row 1)
            for col_num in range(1, len(required_columns) + 1):
                cell = worksheet.cell(row=1, column=col_num)
                cell.border = thin_border
                cell.font = header_font
                cell.alignment = header_alignment
            
            # Styling cho data rows
            for row_num in range(2, len(df) + 2):  # Excel rows start from 1, data starts from row 2
                for col_num in range(1, len(required_columns) + 1):
                    cell = worksheet.cell(row=row_num, column=col_num)
                    cell.border = thin_border
                    cell.alignment = cell_alignment
            
            # Auto-adjust column widths based on content
            for col_num, column in enumerate(required_columns, 1):
                column_letter = get_column_letter(col_num)
                
                # Tính toán width dựa trên header và data
                max_length = len(column)  # Header length
                
                # Kiểm tra độ dài của data trong cột này
                for row_num in range(2, len(df) + 2):
                    cell_value = worksheet.cell(row=row_num, column=col_num).value
                    if cell_value:
                        # Chuyển đổi thành string và tính độ dài
                        cell_length = len(str(cell_value))
                        max_length = max(max_length, cell_length)
                
                # Đặt width với một số padding
                adjusted_width = min(max_length + 2, 50)  # Giới hạn tối đa 50
                worksheet.column_dimensions[column_letter].width = adjusted_width
        
        output.seek(0)
        return output
    
    def create_zip_archive(self, excel_files: Dict[str, io.BytesIO]) -> io.BytesIO:
        """
        Tạo file ZIP chứa tất cả Excel files
        """
        zip_buffer = io.BytesIO()
        
        with zipfile.ZipFile(zip_buffer, 'w', zipfile.ZIP_DEFLATED) as zip_file:
            for filename, excel_data in excel_files.items():
                zip_file.writestr(filename, excel_data.getvalue())
        
        zip_buffer.seek(0)
        return zip_buffer
    
    def export_meta_receipts(
        self, 
        account_ids: List[int], 
        from_date: str, 
        to_date: str
    ) -> io.BytesIO:
        """
        Xuất Meta receipts theo yêu cầu và trả về file ZIP
        """
        # Lấy meta receipts theo điều kiện
        receipts_by_account = self.get_meta_receipt_emails(account_ids, from_date, to_date)
        
        excel_files = {}
        
        for account_id in account_ids:
            # Lấy thông tin account
            account = self.db.query(Account).filter(Account.id == account_id).first()
            if not account:
                continue
            
            account_receipts = receipts_by_account.get(account_id, [])
            
            if account_receipts:  # Chỉ tạo file nếu có data
                # Tạo Excel file cho account này
                excel_data = self.create_excel_file(account_receipts, account.email)
                
                # Tạo tên file: meta_receipts_{email}_{from_date}_{to_date}.xlsx
                safe_email = account.email.replace('@', '_at_').replace('.', '_')
                filename = f"meta_receipts_{safe_email}_{from_date}_{to_date}.xlsx"
                
                excel_files[filename] = excel_data
        
        # Tạo ZIP file
        if excel_files:
            zip_buffer = self.create_zip_archive(excel_files)
            return zip_buffer
        else:
            # Nếu không có data, tạo file ZIP rỗng
            zip_buffer = io.BytesIO()
            with zipfile.ZipFile(zip_buffer, 'w', zipfile.ZIP_DEFLATED) as zip_file:
                zip_file.writestr("no_data.txt", "Không có dữ liệu Meta receipts trong khoảng thời gian này")
            zip_buffer.seek(0)
            return zip_buffer 